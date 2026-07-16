"""Generate downloadable Excel template for call metadata."""

from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


def generate_excel_template() -> BytesIO:
    """Generate a blank Excel template with headers and sample row."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calls"

    # Define headers
    headers = [
        "dietician_name",
        "dietician_id",
        "patient_id",
        "patient_name",
        "appointment_id",
        "call_datetime",
        "recording_url",
        "call_duration_seconds",
    ]

    # Style header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Add example row with format hints
    example_data = [
        "Dr. Rajesh Kumar",                          # dietician_name
        "DTN001",                                     # dietician_id (optional)
        "PAT12345",                                   # patient_id
        "Rajiv Singh",                                # patient_name (optional)
        "APT202401150930",                            # appointment_id
        "2024-01-15 09:30:00",                        # call_datetime (ISO or Excel format)
        "call_001.wav",                               # recording_url (filename or full path)
        "1245",                                       # call_duration_seconds (optional)
    ]

    example_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for col_num, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = value
        cell.fill = example_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border

    # Set column widths
    column_widths = [18, 15, 15, 18, 20, 25, 40, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

    # Add instructions sheet
    instructions_ws = wb.create_sheet("Instructions")
    instructions = [
        ("Column Reference", None),
        ("", None),
        ("dietician_name", "Required. Name of the dietician conducting the call."),
        ("dietician_id", "Optional. Unique ID for the dietician. If provided, takes precedence over name matching."),
        ("patient_id", "Required. Unique patient identifier."),
        ("patient_name", "Optional. Patient name for display."),
        ("appointment_id", "Required. Unique appointment ID. Used for de-duplication — no duplicate appointment IDs will be processed."),
        ("call_datetime", "Required. Call date/time in ISO format (2024-01-15 09:30:00) or Excel date format."),
        ("recording_url", "Required. Audio file reference: filename (e.g., call_001.wav) for local files or HTTP/HTTPS URL for remote files."),
        ("call_duration_seconds", "Optional. Call duration in seconds. If omitted, derived from transcription."),
        ("", None),
        ("Upload Methods", None),
        ("", None),
        ("Method 1: Excel Only (with URLs)", "Use /api/calls/bulk-upload with recording_url as HTTP/HTTPS links to audio files."),
        ("Method 2: Excel + Audio Files", "Use /api/calls/upload-with-audio with Excel file + separate audio files. recording_url should be filename (e.g., 'call_001.wav')."),
        ("", None),
        ("Guidelines", None),
        ("", None),
        ("• Supported audio formats: MP3, WAV, FLAC, M4A, OGG, WebM.", None),
        ("• For Method 1: URLs must be direct links to audio files reachable via HTTP/HTTPS.", None),
        ("• For Method 2: Audio files are matched by filename in the recording_url column.", None),
        ("• Whisper will transcribe all audio files automatically.", None),
        ("• Claude CLI will analyze each call and generate: rubric scores, QA flags, feedback, and metrics.", None),
        ("• Invalid rows will be reported in the validation response before processing begins.", None),
        ("• Once a call is processed, re-uploading the same appointment_id will be rejected.", None),
        ("• Missing optional columns are acceptable; required columns must be present.", None),
        ("• Column order does not matter, but header names must match exactly (case-insensitive, spaces/underscores flexible).", None),
    ]

    for row_num, (title, desc) in enumerate(instructions, 1):
        cell_a = instructions_ws.cell(row=row_num, column=1)
        cell_a.value = title

        if desc:
            cell_b = instructions_ws.cell(row=row_num, column=2)
            cell_b.value = desc
            cell_b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        if title and not desc:
            cell_a.font = Font(bold=True, size=12)
            cell_a.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    instructions_ws.column_dimensions["A"].width = 25
    instructions_ws.column_dimensions["B"].width = 70

    # Save to bytes buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output
