#!/usr/bin/env python3
import requests
import io
from openpyxl import Workbook

# Create a test Excel file
wb = Workbook()
ws = wb.active

# Header
ws['A1'] = 'dietician_name'
ws['B1'] = 'patient_name'
ws['C1'] = 'patient_id'
ws['D1'] = 'appointment_id'
ws['E1'] = 'call_datetime'
ws['F1'] = 'recording_url'

# Test data
ws['A2'] = 'Dr. Priya'
ws['B2'] = 'Ramesh Kumar'
ws['C2'] = 'P001'
ws['D2'] = 'APT-2026-001'
ws['E2'] = '2026-07-07 14:00:00'
ws['F2'] = 'https://dashboard.hellotubelight.com/recording//bajajfinservt//2026-06/6b7898ac-42fc-44e9-8328-8cec7d5e43ad.mp3'

# Save to bytes
output = io.BytesIO()
wb.save(output)
output.seek(0)

# Upload
files = {'file': ('test_upload.xlsx', output, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
response = requests.post('http://localhost:3000/api/calls/bulk-upload', files=files)

if response.status_code == 200:
    result = response.json()
    print(f"Batch ID: {result['batch_id']}")
    print(f"Valid rows: {result['valid_rows']}")
else:
    print(f"Error: {response.status_code}")
    print(response.text)
