#!/usr/bin/env python3
import requests
import io
import json
import time
from openpyxl import Workbook

print("="*80)
print("FULL SYSTEM TEST")
print("="*80)

# Step 1: Create test Excel with 3 calls
print("\n1. Creating test Excel file...")
wb = Workbook()
ws = wb.active
ws['A1'] = 'dietician_name'
ws['B1'] = 'patient_name'
ws['C1'] = 'patient_id'
ws['D1'] = 'appointment_id'
ws['E1'] = 'call_datetime'
ws['F1'] = 'recording_url'

test_calls = [
    ('Dr. Priya', 'Patient A', 'P001', 'APT-001', '2026-07-07 10:00:00',
     'https://dashboard.hellotubelight.com/recording//bajajfinservt//2026-06/6b7898ac-42fc-44e9-8328-8cec7d5e43ad.mp3'),
    ('Dr. Arjun', 'Patient B', 'P002', 'APT-002', '2026-07-07 11:00:00',
     'https://dashboard.hellotubelight.com/recording//bajajfinservt//2026-06/6b7898ac-42fc-44e9-8328-8cec7d5e43ad.mp3'),
    ('Dr. Neha', 'Patient C', 'P003', 'APT-003', '2026-07-07 12:00:00',
     'https://dashboard.hellotubelight.com/recording//bajajfinservt//2026-06/6b7898ac-42fc-44e9-8328-8cec7d5e43ad.mp3'),
]

for idx, (d, p, pid, aid, dt, url) in enumerate(test_calls, 2):
    ws[f'A{idx}'] = d
    ws[f'B{idx}'] = p
    ws[f'C{idx}'] = pid
    ws[f'D{idx}'] = aid
    ws[f'E{idx}'] = dt
    ws[f'F{idx}'] = url

output = io.BytesIO()
wb.save(output)
output.seek(0)
print("   Created with 3 test calls")

# Step 2: Upload
print("\n2. Uploading to server...")
files = {'file': ('test.xlsx', output, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
response = requests.post('http://localhost:8000/api/calls/bulk-upload', files=files)

if response.status_code != 200:
    print(f"   ERROR: {response.status_code}")
    print(response.text)
    exit(1)

report = response.json()
batch_id = report['batch_id']
print(f"   Batch ID: {batch_id}")
print(f"   Valid: {report['valid_rows']}, Invalid: {report['invalid_rows']}")

# Step 3: Wait for processing
print("\n3. Waiting for processing to complete...")
for attempt in range(60):
    response = requests.get(f'http://localhost:8000/api/batches/{batch_id}/progress')
    if response.status_code == 200:
        progress = response.json()
        pct = progress.get('pct_complete', 0)
        print(f"   Progress: {pct}%")
        if pct == 100:
            print("   COMPLETE!")
            break
    time.sleep(2)

# Step 4: Check results
print("\n4. Checking results...")
response = requests.get('http://localhost:8000/api/calls')
calls = response.json()

if not calls:
    print("   ERROR: No calls found!")
    exit(1)

print(f"   Found {len(calls)} calls")

for call in calls[:3]:
    score = call.get('overall_weighted_score')
    print(f"\n   Call: {call['dietician_name']}")
    print(f"     Score: {score}")

print("\n" + "="*80)
print("TEST COMPLETE - Check portal at http://localhost:8000")
print("="*80)
